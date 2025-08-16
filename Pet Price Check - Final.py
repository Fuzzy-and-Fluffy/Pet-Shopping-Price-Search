import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, WebDriverException
import pandas as pd
from urllib.parse import urlparse
import time
import logging
import traceback
import random
import openpyxl
import datetime

#版本信息：抓取各项数据ok，店铺名称完全抓取成功，小规模测试可,可以抓取cat evolution,但是output excel格式和公式未保留

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)
    options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    
    driver = webdriver.Chrome(options=options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver

def get_99petshops_search_results(driver, url):
    try:
        logger.info(f"正在处理99petshops搜索页面: {url}")
        driver.get(url)
        time.sleep(10)  # 等待页面完全加载
        
        price = None
        store = None
        variant_count = 0

        # 尝试找到所有产品卡片
        product_cards = driver.find_elements(By.CSS_SELECTOR, ".pd-info, .product-card")
        
        if not product_cards:
            logger.warning("没有找到产品卡片")
            return None, None, 0
            
        # 计算变体数量 - 改进的方法
        variant_count = len(product_cards)
        unique_titles = set()
        
        for card in product_cards:
            try:
                title = card.find_element(By.CSS_SELECTOR, "h2, .product-title").text.strip()
                unique_titles.add(title)
            except NoSuchElementException:
                continue
                
        variant_count = len(unique_titles)
        logger.info(f"找到 {variant_count} 个独特产品变体")
        
        # 尝试多种方式获取价格和商店信息
        try:
            # 方法1: 寻找高亮显示的价格
            highlighted = driver.find_element(By.CSS_SELECTOR, "span.hilighted")
            price_element = highlighted.find_element(By.CSS_SELECTOR, "span")
            price_text = price_element.text
            
            # 从价格中提取数字
            price_match = re.search(r'\$?(\d+\.?\d*)', price_text)
            if price_match:
                price = float(price_match.group(1))
                logger.info(f"从hilighted找到价格: {price}")
                
                # 获取商店信息
                try:
                    store_img = highlighted.find_element(By.XPATH, "./following-sibling::img")
                    store = store_img.get_attribute('alt')
                    logger.info(f"从hilighted找到商店: {store}")
                except NoSuchElementException:
                    logger.warning("没有找到商店信息的img标签")
        
        except NoSuchElementException:
            # 方法2: 寻找普通价格
            try:
                normal_price = driver.find_element(By.CSS_SELECTOR, "span.normal")
                price_element = normal_price.find_element(By.CSS_SELECTOR, "span")
                price_text = price_element.text
                price_match = re.search(r'\$?(\d+\.?\d*)', price_text)
                if price_match:
                    price = float(price_match.group(1))
                    logger.info(f"从normal找到价格: {price}")
                    
                    # 获取商店信息
                    try:
                        store_img = normal_price.find_element(By.XPATH, "./following-sibling::img")
                        store = store_img.get_attribute('alt')
                        logger.info(f"从normal找到商店: {store}")
                    except NoSuchElementException:
                        logger.warning("没有找到商店信息的img标签")
            except NoSuchElementException:
                logger.warning("没有找到普通价格信息")

        return price, store, variant_count if variant_count > 0 else 1  # 确保变体数至少为1

    except Exception as e:
        logger.error(f"处理搜索页面时发生错误: {str(e)}")
        logger.error(traceback.format_exc())
        return None, None, None


def get_catevolution_price(driver, url):
    try:
        logger.info(f"正在处理Catevolution页面: {url}")
        driver.get(url)
        time.sleep(5)  # 等待页面完全加载

        price = None
        store = "Catevolution"
        variant_count = 1  # 固定为1，因为是单个产品页面

        # 尝试从meta标签中获取价格信息
        try:
            price_meta = driver.find_element(By.CSS_SELECTOR, "meta[property='og:price:amount']")
            price_text = price_meta.get_attribute('content')
            if price_text:
                price = float(price_text)
                logger.info(f"从meta标签找到价格: {price}")
        except NoSuchElementException:
            logger.warning("没有找到价格信息的meta标签")

        return price, store, variant_count

    except Exception as e:
        logger.error(f"处理Catevolution页面时发生错误: {str(e)}")
        logger.error(traceback.format_exc())
        return None, None, None

def main():
    logger.info("开始运行爬虫程序")

    try:
        # 使用 openpyxl 加载 Excel 文件
        workbook = openpyxl.load_workbook("Pet Shopping List.xlsx")
        sheet = workbook.active  # 获取活动的工作表

        logger.info(f"成功读取Excel文件，共{sheet.max_row - 1}行数据")

        # 检查必要的列是否存在
        required_columns = ['链接', '最低价', '最低价商店', 'Product Variants']
        header_row = [cell.value for cell in sheet[1]]  # 假设第一行是标题行
        if not all(col in header_row for col in required_columns):
            logger.error("Excel文件缺少必要的列")
            return

        # 获取列索引
        column_index = {col: header_row.index(col) + 1 for col in required_columns}

        driver = setup_driver()
        logger.info("成功初始化WebDriver")

        try:
            for row in range(2, sheet.max_row + 1):  # 从第二行开始
                url = sheet.cell(row=row, column=column_index['链接']).value
                if not url:
                    logger.warning(f"第{row}行URL为空，跳过")
                    continue

                logger.info(f"正在处理第{row}行: {url}")

                if '99petshops.com.au' in url:
                    price, store, variant_count = get_99petshops_search_results(driver, url)
                elif 'catevolution.com.au' in url:
                    price, store, variant_count = get_catevolution_price(driver, url)
                else:
                    logger.warning(f"不支持的URL: {url}")
                    continue

                if price is not None:
                    sheet.cell(row=row, column=column_index['最低价'], value=price)
                if store is not None:
                    sheet.cell(row=row, column=column_index['最低价商店'], value=store)
                if variant_count is not None:
                    sheet.cell(row=row, column=column_index['Product Variants'], value=variant_count)

                logger.info(f"第{row}行数据更新完成: 价格={price}, 商店={store}, 变体数量={variant_count}")

                time.sleep(2 + random.random() * 3)

            # 在保存文件时，动态生成文件名
            current_time = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")  # 格式化日期时间为 "YYYYMMDD_HHMMSS"
            output_filename = f"Pet Shopping List_updated_{current_time}.xlsx"
            workbook.save(output_filename)
            logger.info(f"所有数据处理完成，已保存到新的Excel文件: {output_filename}")

        finally:
            driver.quit()
            logger.info("已关闭WebDriver")

    except Exception as e:
        logger.error(f"程序运行出错: {str(e)}")
        logger.error(traceback.format_exc())

if __name__ == "__main__":
    main()




